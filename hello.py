from bs4 import BeautifulSoup
from urllib import request
import openpyxl
list_name = []
list_score = []
list_episode_latest = []
url = "https://animehay.club/loc-phim/W1tdLFsyMDIxXSxbXSxbXV0="
req = request.Request( url=url, headers={'User-Agent': 'Mozilla/5.0'} )
page = request.urlopen(req)
soup = BeautifulSoup(page, 'html.parser')
movie_list = soup.findAll('div', class_='movie-item')
for movie in movie_list:
    movie_name = movie.find('a').get('title')
    list_name.append(movie_name)
    movie_score = movie.find('div', class_='score').text
    list_score.append(movie_score)
    movie_episode_latest = movie.find('div', class_='episode-latest').text
    list_episode_latest.append(movie_episode_latest)
wb = openpyxl.Workbook()
ws = wb.active
ws.cell(column = 1, row = 1, value = "name_film")
ws.cell(column = 2, row = 1, value = "score")
ws.cell(column = 3, row = 1, value = "episode-latest")
for i in range(0, len(list_name)):
    ws.cell(column =1, row = i + 2, value=list_name[i])
    ws.cell(column = 2, row = i + 2, value=list_score[i])
    ws.cell(column = 3, row = i + 2, value=list_episode_latest[i])
wb.save('ouput.xlsx')
